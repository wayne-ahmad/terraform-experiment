import logging

from openpyxl import load_workbook
from dataquery_processor import get_config_path
import csv
import os
from datetime import date


def rotate_list(list_1, n):
    """
    Rotates a list n positions. Used to push measure to the front of the data
    :param list_1: the list to rotate
    :param n: number of positions to rotate
    :return: rotated list
    """
    return (list_1[-n:] + list_1[:-n])


class ExcelHandler(object):
    """
    Class for creating Excel outputs
    """

    def __init__(self, output_file=None, csv_file=None, manifest=None):
        """
        Set up the handler

        :param output_file: the output path for the Excel pivot
        :param csv_file: the CSV data file to use as the basis for the Excel output
        :param manifest: the request manifest that includes column metadata
        """
        self.csv_file = csv_file
        self.filepath = output_file
        self.template = get_config_path("excel_template.xlsx")
        self.workbook = None
        self.rows = None
        self.cols = None
        self.manifest = manifest
        self.total = 0
        self.__create_blank_workbook_from_template__()

    def create_notes_only(self, filename=None):
        """
        Create a Notes page only, to accompany CSV outputs
        :param filename: the path/filename to save the notes file
        :return:
        """
        self.__update_notes__(format='CSV')
        self.workbook.remove(self.workbook.get_sheet_by_name('Data'))
        self.workbook.remove(self.workbook.get_sheet_by_name('Pivot'))
        if filename is not None:
            self.workbook.save(filename=filename)
        else:
            self.workbook.save(filename=self.filepath)

    def create_workbook(self):
        """
        Creates a workbook from a CSV file, including pivot table and notes
        :return:
        """
        if self.csv_file is not None:
            self.__populate_worksheet_from_csv_file__()
            self.__update_pivot_table__()
            self.__update_notes__()
            self.workbook.save(filename=self.filepath)

    def __create_blank_workbook_from_template__(self):
        """
        Creates a fresh workbook from a template
        :return: None
        """
        self.workbook = load_workbook(self.template)
        sheet_name = 'Data'
        idx = self.workbook.sheetnames.index(sheet_name)
        ws = self.workbook.get_sheet_by_name(sheet_name)
        self.workbook.remove(ws)
        self.workbook.create_sheet(sheet_name, idx)

    def __populate_worksheet_from_csv_file__(self):
        """
        Populates a workbook using CSV file data
        :return: None
        """
        if not os.path.exists(self.csv_file):
            raise ValueError("There is no CSV file to export")
        ws = self.workbook.get_sheet_by_name('Data')
        with open(self.csv_file, encoding='utf-8') as f:
            reader = csv.reader(f, quoting=csv.QUOTE_NONNUMERIC)
            # Header
            ws.append(rotate_list(reader.__next__(), 1))
            for row in reader:
                if isinstance(row, list):
                    ws.append(rotate_list(row, 1))
                    try:
                        self.total += int(row[len(row)-1])
                    except ValueError:
                        logging.warning("There is a row with no measure")
        self.rows = reader.line_num
        self.cols = len(row)

    def __populate_worksheet_from_odbc_cursor__(self, cursor=None, headers=None):
        """
        Populates workbook from a PyODBC cursor
        :param cursor: the PyODBC cursor
        :param headers: a list of headers
        :return: None
        """
        self.rows = 0
        self.cols = len(headers)
        ws = self.workbook.get_sheet_by_name('Data')
        # Add header
        ws.append(rotate_list(headers, 1))
        for row in cursor.fetchall():
            row_list = [elem for elem in row]
            ws.append(rotate_list(row_list, 1))
            self.rows += 1
            self.total += int(row[len(row) - 1])
        self.__update_pivot_table__()
        self.__update_notes__()
        self.workbook.save(filename=self.filepath)

    def __update_pivot_table__(self):
        ws = self.workbook["Pivot"]
        pivot = ws._pivots[0]
        # Update the pivot table range
        pivot.cache.cacheSource.worksheetSource.ref = self.__range__()
        # Set to refresh
        pivot.cache.refreshOnLoad = True

    def __update_notes__(self, format='Excel pivot table'):
        ws = self.workbook["Notes"]
        ws['B6'] = self.manifest['orderRef']
        ws['B7'] = self.manifest['customerRef']
        ws['B9'] = 'Jisc Tailored Datasets App v1.0'
        ws['B11'] = date.today().isoformat()
        ws['B14'] = self.manifest['datasource']
        ws['B15'] = self.total
        ws['B16'] = self.manifest['onwardUseCategory']
        ws['B17'] = format

        # Field names and definitions
        row = 32
        for item in self.manifest['items']:
            if 'allowedValues' not in item.keys() or len(item['allowedValues']) == 0:
                ws['A' + str(row)] = item['fieldName']
                if 'description' in item.keys():
                    ws['B' + str(row)] = item['description']
                row += 1

    def __range__(self):
        """
        Determine the cell range based on the dataset size
        :return: a cell range e.g "A1:C26"
        """
        return 'A1' + ':' + chr(ord('@') + self.cols) + str(self.rows)